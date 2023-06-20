from types import TracebackType
from typing import Callable, Any

from openpyxl import load_workbook
from Transformations.Classes import Slide
from Transformations.Mapping import SLIDE_ID, CONTENT_LOCATION, CONTENT


SLIDES_TO_POPULATE = {}


def parse_excel(filePath: str) -> dict[Any, Any] | Callable[[TracebackType | None], Any]:
    """
    The parse excel function parses the active workbook (we default to the first sheet in the workbook).

    Current functionality supported is limited to two columns, placement Id and content.
    We populate a dictionary of PlacementID:Content and return it to our main function.
    """
    try:
        workbook = load_workbook(filePath)
        sheet = workbook.active

        print(f"Parsing workbook '{sheet.title}'")
        # iterate over all rows in the workbook
        # we use min_row=2 to skip the header rows.
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # if the slide id is in the slides_to_populate, add to it.
            if row[SLIDE_ID] in SLIDES_TO_POPULATE.keys():
                SLIDES_TO_POPULATE[row[SLIDE_ID]].add(row[CONTENT_LOCATION], row[CONTENT])
            else:
                slide = Slide(row[SLIDE_ID])
                slide.add(row[CONTENT_LOCATION], row[CONTENT])
                SLIDES_TO_POPULATE[slide.slide_id] = slide
        return SLIDES_TO_POPULATE

    except FileNotFoundError:
        return FileNotFoundError.with_traceback
